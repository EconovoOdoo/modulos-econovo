# -*- coding: utf-8 -*-
"""HTTP controller: export BOM Cost Summary to Excel (.xlsx).

Endpoint: GET /econovo/bom_cost_summary/export_xlsx
Params:   bom_id, quantity, variant, warehouse_id, costs, operations, lead_times

Returns a .xlsx workbook with three sheets:
  Sheet 1 "BOM Tree"           — hierarchical BOM breakdown mirroring the
                                  native Odoo BOM Overview tree view, with
                                  expandable operations and byproducts per level.
  Sheet 2 "Cost Summary"       — costs aggregated by product category / work center.
  Sheet 3 "Components Detail"  — flat pivot-ready list, one row per usage.
"""

import io
import logging

from odoo import http
from odoo.http import content_disposition, request

_logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
_C = {
    "header_bg": "1F4E79",  # section header background (dark blue)
    "header_fg": "FFFFFF",  # section header foreground (white)
    "info_bg":   "D9D9D9",  # BOM info rows
    "cat_0":     "BDD7EE",  # category depth 0
    "cat_1":     "DDEBF7",  # category depth 1
    "cat_deep":  "EEF4FB",  # category depth 2+
    "product":   "F2F2F2",  # product rows
    "usage":     "FFFFFF",  # usage rows
    "wc":        "E2EFDA",  # workcenter rows
    "op":        "F4F9F5",  # operation rows
    "subtotal":  "FFF2CC",  # subtotal rows
    "total":     "FFE082",  # grand total row
    # Subcontracting section colours (purple tones)
    "sc_vendor":   "E8D5F5",  # subcontracting vendor group row
    "sc_item":     "F5EEF8",  # subcontracting item row
    "sc_subtotal": "D7BDE2",  # subcontracting section subtotal row
    # Byproduct section colours (green tones — value recovered from the process)
    "bp_cat_0":    "C6E0B4",  # byproduct category depth 0
    "bp_cat_1":    "D9EAD3",  # byproduct category depth 1
    "bp_cat_deep": "EAF4E7",  # byproduct category depth 2+
    "bp_product":  "F3FAF1",  # byproduct product rows
    "bp_subtotal": "A9D18E",  # byproduct section subtotal row
    # BOM Tree sheet
    "tree_root":    "1F4E79",  # root product row (dark blue)
    "tree_root_fg": "FFFFFF",  # root product foreground (white)
    "tree_bom":     "BDD7EE",  # sub-BOM rows (medium blue)
    "tree_leaf":    "F2F2F2",  # leaf component rows
    "tree_ops":     "E2EFDA",  # operations group header
    "tree_op":      "F4F9F5",  # individual operation rows
    "tree_bp":      "A9D18E",  # byproducts group header
    "tree_bp_item": "D9EAD3",  # individual byproduct rows
    "tree_sc":      "E8D5F5",  # subcontracting row
}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    """Solid fill from a hex colour string."""
    from openpyxl.styles import PatternFill  # noqa: PLC0415
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9):
    from openpyxl.styles import Font  # noqa: PLC0415
    return Font(bold=bold, color=color, size=size)


def _align_right():
    from openpyxl.styles import Alignment  # noqa: PLC0415
    return Alignment(horizontal="right", vertical="center")


def _align_left():
    from openpyxl.styles import Alignment  # noqa: PLC0415
    return Alignment(horizontal="left", vertical="center")


def _align_center():
    from openpyxl.styles import Alignment  # noqa: PLC0415
    return Alignment(horizontal="center", vertical="center")


# ── Value helpers ─────────────────────────────────────────────────────────────

def _flt(value):
    """Return a rounded float or empty string for False/None."""
    if value is None or value is False:
        return ""
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return ""


def _pct(value):
    """Return e.g. '45.2%' or '' for missing values."""
    if value is None or value is False:
        return ""
    try:
        return "%.1f%%" % float(value)
    except (TypeError, ValueError):
        return ""


def _str(value):
    """Return value as string, empty string for None/False."""
    if value is None or value is False:
        return ""
    return str(value)


# ── Low-level row writer ──────────────────────────────────────────────────────

def _write_row(ws, row_idx, values, bg_color, bold=False, outline_level=0):
    """Write a data row, applying background colour and optional outline."""
    fill = _fill(bg_color)
    row_font = _font(bold=bold)
    # Right-align numeric columns (indices 4-onward, but only if the value
    # is a float/int; text columns stay left-aligned).
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = row_font
        cell.fill = fill
        if isinstance(val, (int, float)) and col_idx > 3:
            cell.alignment = _align_right()
        else:
            cell.alignment = _align_left()
    # Cap at Excel's maximum outline depth of 7.
    if outline_level:
        ws.row_dimensions[row_idx].outline_level = min(outline_level, 7)


def _write_info(ws, row_idx, label, value, col_count):
    """Write a BOM-info banner row (label in col A, value in col B)."""
    fill = _fill(_C["info_bg"])
    c = ws.cell(row=row_idx, column=1, value=label)
    c.font = _font(bold=True)
    c.fill = fill
    c.alignment = _align_left()
    c2 = ws.cell(row=row_idx, column=2, value=value)
    c2.font = _font()
    c2.fill = fill
    c2.alignment = _align_left()
    for col_idx in range(3, col_count + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _write_section_header(ws, row_idx, title, col_count):
    """Write a full-width section header row."""
    fill = _fill(_C["header_bg"])
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
    c = ws.cell(row=row_idx, column=1, value=title)
    c.font = _font(bold=True, color=_C["header_fg"], size=10)
    c.alignment = _align_left()
    ws.row_dimensions[row_idx].height = 16


def _cell_comment(ws, row, col, text, author="Econovo"):
    """Attach a sticky-note comment to a worksheet cell (visible on hover)."""
    from openpyxl.comments import Comment  # noqa: PLC0415
    note = Comment(text, author)
    note.width = 380
    note.height = 160
    ws.cell(row=row, column=col).comment = note


def _write_subtotal(ws, row_idx, label, col_count, ci,
                    bom_cost, bom_cost_usd, prod_cost, prod_cost_usd,
                    show_costs, has_usd, cur, usd, color=None):
    """Write a subtotal row."""
    vals = [""] * col_count
    vals[ci["Type"] - 1] = "Subtotal"
    vals[ci["Name"] - 1] = label
    if show_costs:
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(bom_cost)
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(bom_cost_usd)
        if prod_cost is not None and "Product Cost (%s)" % cur in ci:
            vals[ci["Product Cost (%s)" % cur] - 1] = _flt(prod_cost)
        if prod_cost_usd is not None and has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(prod_cost_usd)
    _write_row(ws, row_idx, vals, color or _C["subtotal"], bold=True, outline_level=0)
    return row_idx + 1


# ── Recursive category writer (Sheet 1) ──────────────────────────────────────

def _write_category_rows(ws, row, node, ci, cur, usd,
                         has_usd, show_costs, show_lead_times,
                         outline_base):
    """
    Recursively write category → product → usage rows for one tree node.

    :param outline_base: outline_level assigned to this category row.
                         Its products get outline_base+1, usages outline_base+2.
    """
    col_count = len(ci)
    depth = node.get("depth", 0)
    cat_colors = [_C["cat_0"], _C["cat_1"], _C["cat_deep"]]
    cat_color = cat_colors[min(depth, len(cat_colors) - 1)]
    indent = "    " * depth

    # Category row
    vals = [""] * col_count
    vals[ci["Level"] - 1] = depth
    vals[ci["Type"] - 1] = "Category"
    vals[ci["Name"] - 1] = indent + node["name"]
    vals[ci["%"] - 1] = _pct(node.get("percentage"))
    if show_costs:
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(node.get("total"))
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(node.get("total_usd"))
        if "Product Cost (%s)" % cur in ci:
            vals[ci["Product Cost (%s)" % cur] - 1] = _flt(node.get("prod_cost_total"))
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(node.get("prod_cost_total_usd"))
    _write_row(ws, row, vals, cat_color, bold=True, outline_level=outline_base)
    row += 1

    # Recurse into child categories first
    for child in node.get("children", []):
        row = _write_category_rows(
            ws, row, child, ci, cur, usd,
            has_usd, show_costs, show_lead_times,
            outline_base=outline_base + 1,
        )

    # Products under this category
    for prod in node.get("products", []):
        usages = prod.get("usages", [])

        # Aggregate quantity (only if single UoM across all usages)
        uoms = {u.get("uom_name", "") for u in usages if u.get("uom_name")}
        if len(uoms) == 1:
            total_qty = _flt(sum(u.get("quantity", 0) for u in usages))
            uom_display = list(uoms)[0]
        elif len(uoms) > 1:
            total_qty = ""        # mixed UoMs → cannot aggregate
            uom_display = "—"
        else:
            total_qty = ""
            uom_display = ""

        prod_indent = "    " * (depth + 1)
        vals = [""] * col_count
        vals[ci["Level"] - 1] = depth + 1
        vals[ci["Type"] - 1] = "Product"
        vals[ci["Name"] - 1] = prod_indent + prod["name"]
        vals[ci["Qty / Duration"] - 1] = total_qty
        vals[ci["UoM"] - 1] = uom_display
        vals[ci["%"] - 1] = _pct(prod.get("percentage"))
        if show_costs:
            vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(prod.get("total"))
            if has_usd and "BOM Cost (%s)" % usd in ci:
                vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(prod.get("total_usd"))
            if "Product Cost (%s)" % cur in ci:
                vals[ci["Product Cost (%s)" % cur] - 1] = _flt(prod.get("prod_cost_total"))
            if has_usd and "Product Cost (%s)" % usd in ci:
                vals[ci["Product Cost (%s)" % usd] - 1] = _flt(prod.get("prod_cost_total_usd"))
        # Availability (always present columns)
        qa = prod.get("quantity_available")
        if qa is not False and qa is not None and "Free to Use" in ci:
            vals[ci["Free to Use"] - 1] = _flt(qa)
            vals[ci["On Hand"] - 1] = _flt(prod.get("quantity_on_hand"))
        if "Availability" in ci:
            vals[ci["Availability"] - 1] = _str(prod.get("availability_display"))

        _write_row(ws, row, vals, _C["product"], bold=False,
                   outline_level=outline_base + 1)
        row += 1

        # Usage rows
        for usage in usages:
            usage_indent = "    " * (depth + 2)
            parent_name = _str(usage.get("parent_name")) or "—"
            vals = [""] * col_count
            vals[ci["Level"] - 1] = depth + 2
            vals[ci["Type"] - 1] = "Usage"
            vals[ci["Name"] - 1] = usage_indent + parent_name
            vals[ci["Qty / Duration"] - 1] = _flt(usage.get("quantity"))
            vals[ci["UoM"] - 1] = _str(usage.get("uom_name"))
            vals[ci["%"] - 1] = _pct(usage.get("percentage"))
            if show_costs:
                vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(usage.get("total"))
                if has_usd and "BOM Cost (%s)" % usd in ci:
                    vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(usage.get("total_usd"))
                if "Product Cost (%s)" % cur in ci:
                    vals[ci["Product Cost (%s)" % cur] - 1] = _flt(usage.get("prod_cost"))
                if has_usd and "Product Cost (%s)" % usd in ci:
                    vals[ci["Product Cost (%s)" % usd] - 1] = _flt(usage.get("prod_cost_usd"))
            if show_lead_times:
                lt = usage.get("lead_time")
                vals[ci["Lead Time (days)"] - 1] = lt if lt is not False and lt is not None else ""
                route = _str(usage.get("route_name"))
                detail = _str(usage.get("route_detail"))
                if route and detail:
                    route = route + ": " + detail
                elif detail:
                    route = detail
                vals[ci["Route"] - 1] = route
            _write_row(ws, row, vals, _C["usage"], outline_level=outline_base + 2)
            row += 1

    return row


# ── Recursive byproduct category writer (Sheet 1) ─────────────────────────────

def _write_byproduct_category_rows(ws, row, node, ci, cur, usd,
                                   has_usd, show_costs, outline_base):
    """
    Recursively write byproduct category → product → usage rows.

    Byproducts show two percentages stacked in the ``%`` column:
      * BOM-cost-based % (percentage) — may be 0 when cost_share = 0%
      * Prod-cost-based % (prod_cost_percentage) — always meaningful

    BOM Cost column  = amount allocated away from main product (bom_cost).
    Product Cost col = recoverable standard catalogue value (prod_cost).
    """
    col_count = len(ci)
    depth = node.get("depth", 0)
    cat_colors = [_C["bp_cat_0"], _C["bp_cat_1"], _C["bp_cat_deep"]]
    cat_color = cat_colors[min(depth, len(cat_colors) - 1)]
    indent = "    " * depth

    # Category row
    pct_bom = node.get("percentage", 0)
    pct_val = node.get("prod_cost_percentage", 0)
    pct_display = "%.1f%% / %.1f%%" % (pct_bom, pct_val)
    vals = [""] * col_count
    vals[ci["Level"] - 1] = depth
    vals[ci["Type"] - 1] = "BP Category"
    vals[ci["Name"] - 1] = indent + node["name"]
    vals[ci["%"] - 1] = pct_display
    if show_costs:
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(node.get("total"))
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(node.get("total_usd"))
        if "Product Cost (%s)" % cur in ci:
            vals[ci["Product Cost (%s)" % cur] - 1] = _flt(node.get("prod_cost_total"))
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(node.get("prod_cost_total_usd"))
    _write_row(ws, row, vals, cat_color, bold=True, outline_level=outline_base)
    row += 1

    # Recurse into child categories
    for child in node.get("children", []):
        row = _write_byproduct_category_rows(
            ws, row, child, ci, cur, usd,
            has_usd, show_costs,
            outline_base=outline_base + 1,
        )

    # Byproduct products under this category
    for prod in node.get("products", []):
        usages = prod.get("usages", [])

        uoms = {u.get("uom_name", "") for u in usages if u.get("uom_name")}
        if len(uoms) == 1:
            total_qty = _flt(sum(u.get("quantity", 0) for u in usages))
            uom_display = list(uoms)[0]
        elif len(uoms) > 1:
            total_qty = ""
            uom_display = "—"
        else:
            total_qty = ""
            uom_display = ""

        prod_pct_bom = prod.get("percentage", 0)
        prod_pct_val = prod.get("prod_cost_percentage", 0)
        prod_pct_display = "%.1f%% / %.1f%%" % (prod_pct_bom, prod_pct_val)
        prod_indent = "    " * (depth + 1)
        vals = [""] * col_count
        vals[ci["Level"] - 1] = depth + 1
        vals[ci["Type"] - 1] = "BP Product"
        vals[ci["Name"] - 1] = prod_indent + prod["name"]
        vals[ci["Qty / Duration"] - 1] = total_qty
        vals[ci["UoM"] - 1] = uom_display
        vals[ci["%"] - 1] = prod_pct_display
        if show_costs:
            vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(prod.get("total"))
            if has_usd and "BOM Cost (%s)" % usd in ci:
                vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(prod.get("total_usd"))
            if "Product Cost (%s)" % cur in ci:
                vals[ci["Product Cost (%s)" % cur] - 1] = _flt(prod.get("prod_cost_total"))
            if has_usd and "Product Cost (%s)" % usd in ci:
                vals[ci["Product Cost (%s)" % usd] - 1] = _flt(prod.get("prod_cost_total_usd"))
        _write_row(ws, row, vals, _C["bp_product"], bold=False,
                   outline_level=outline_base + 1)
        row += 1

        # Usage rows
        for usage in usages:
            usage_pct_bom = usage.get("percentage", 0)
            usage_pct_val = usage.get("prod_cost_percentage", 0)
            usage_pct_display = "%.1f%% / %.1f%%" % (usage_pct_bom, usage_pct_val)
            usage_indent = "    " * (depth + 2)
            parent_name = _str(usage.get("parent_name")) or "—"
            vals = [""] * col_count
            vals[ci["Level"] - 1] = depth + 2
            vals[ci["Type"] - 1] = "BP Usage"
            vals[ci["Name"] - 1] = usage_indent + parent_name
            vals[ci["Qty / Duration"] - 1] = _flt(usage.get("quantity"))
            vals[ci["UoM"] - 1] = _str(usage.get("uom_name"))
            vals[ci["%"] - 1] = usage_pct_display
            if show_costs:
                vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(usage.get("total"))
                if has_usd and "BOM Cost (%s)" % usd in ci:
                    vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(usage.get("total_usd"))
                if "Product Cost (%s)" % cur in ci:
                    vals[ci["Product Cost (%s)" % cur] - 1] = _flt(usage.get("prod_cost"))
                if has_usd and "Product Cost (%s)" % usd in ci:
                    vals[ci["Product Cost (%s)" % usd] - 1] = _flt(usage.get("prod_cost_usd"))
            _write_row(ws, row, vals, _C["usage"], outline_level=outline_base + 2)
            row += 1

    return row


# ════════════════════════════ Sheet 1: BOM Tree ══════════════════════════════

def _tw_row(ws, row_idx, vals, bg, outline=0, bold=False, fg="000000"):
    """Write a BOM-tree row with custom foreground colour support."""
    fnt = _font(bold=bold, color=fg)
    fill = _fill(bg)
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = fnt
        cell.fill = fill
        if isinstance(val, (int, float)) and col_idx > 2:
            cell.alignment = _align_right()
        else:
            cell.alignment = _align_left()
    if outline:
        ws.row_dimensions[row_idx].outline_level = min(outline, 7)


def _tree_costs(vals, ci, cur, usd, has_usd, rate, bom_cost, prod_cost):
    """Fill BOM Cost and Product Cost columns for a tree row in-place."""
    if "BOM Cost (%s)" % cur in ci:
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(bom_cost)
    if has_usd and "BOM Cost (%s)" % usd in ci and bom_cost not in (None, False, "") and rate:
        vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(bom_cost * rate)
    if "Product Cost (%s)" % cur in ci:
        vals[ci["Product Cost (%s)" % cur] - 1] = _flt(prod_cost)
    if has_usd and "Product Cost (%s)" % usd in ci and prod_cost not in (None, False, "") and rate:
        vals[ci["Product Cost (%s)" % usd] - 1] = _flt(prod_cost * rate)


def _write_tree_node(ws, row, node, ci, cur, usd, rate,
                     has_usd, show_costs, show_operations, show_lead_times,
                     outline_level, col_count):
    """
    Recursively write one BOM node (root, sub-BOM, or leaf component) and
    all its children, then its operations and byproducts as collapsible groups.

    outline_level 0 = root product row (always visible, dark blue).
    outline_level N = components / ops / byproducts at depth N.
    """  # noqa: E501
    is_root = outline_level == 0
    is_sub_bom = node.get("type") == "bom" and not is_root

    if is_root:
        bg, fg, bold = _C["tree_root"], _C["tree_root_fg"], True
    elif is_sub_bom:
        bg, fg, bold = _C["tree_bom"], "000000", True
    else:
        bg, fg, bold = _C["tree_leaf"], "000000", False

    indent = "    " * outline_level
    vals = [""] * col_count
    vals[ci["Name"] - 1] = indent + _str(node.get("name", ""))
    vals[ci["Qty"] - 1] = _flt(node.get("quantity"))
    vals[ci["UoM"] - 1] = _str(node.get("uom_name"))
    if show_lead_times:
        lt = node.get("lead_time")
        vals[ci["Lead Time (days)"] - 1] = lt if lt not in (None, False) else ""
        route = _str(node.get("route_name", ""))
        detail = _str(node.get("route_detail", ""))
        if route and detail:
            route = route + ": " + detail
        elif detail:
            route = detail
        vals[ci["Route"] - 1] = route
    if show_costs:
        _tree_costs(
            vals, ci, cur, usd, has_usd, rate,
            node.get("bom_cost"), node.get("prod_cost"),
        )
    if not is_sub_bom and not is_root:
        # Availability only for leaf components and their parent sub-BOM rows
        vals[ci["Free to Use"] - 1] = _flt(node.get("quantity_available"))
        vals[ci["On Hand"] - 1] = _flt(node.get("quantity_on_hand"))
        vals[ci["Availability"] - 1] = _str(node.get("availability_display"))
    _tw_row(ws, row, vals, bg, outline=outline_level, bold=bold, fg=fg)
    if is_root:
        ws.row_dimensions[row].height = 16
    row += 1

    # ── Recurse into direct components ───────────────────────────────────────
    for comp in node.get("components", []):
        row = _write_tree_node(
            ws, row, comp, ci, cur, usd, rate,
            has_usd, show_costs, show_operations, show_lead_times,
            outline_level=outline_level + 1,
            col_count=col_count,
        )

    child_outline = outline_level + 1

    # ── Operations group for this BOM level ──────────────────────────────────
    if show_operations:
        ops = node.get("operations", [])
        if ops:
            ops_cost = node.get("operations_cost",
                                sum(o.get("bom_cost", 0) for o in ops))
            ops_time = node.get("operations_time",
                                sum(o.get("quantity", 0) for o in ops))
            ops_vals = [""] * col_count
            ops_vals[ci["Name"] - 1] = (
                "    " * child_outline + "\u25b6 Operations (%d)" % len(ops)
            )
            ops_vals[ci["Qty"] - 1] = _flt(ops_time)
            ops_vals[ci["UoM"] - 1] = "min"
            if show_costs and "BOM Cost (%s)" % cur in ci:
                ops_vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(ops_cost)
                if has_usd and "BOM Cost (%s)" % usd in ci and rate:
                    ops_vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(
                        ops_cost * rate
                    )
            _tw_row(ws, row, ops_vals, _C["tree_ops"],
                    outline=child_outline, bold=True)
            row += 1

            for op in ops:
                op_vals = [""] * col_count
                op_vals[ci["Name"] - 1] = (
                    "    " * (child_outline + 1) + _str(op.get("name", ""))
                )
                op_vals[ci["Qty"] - 1] = _flt(op.get("quantity"))  # minutes
                op_vals[ci["UoM"] - 1] = "min"
                if show_costs and "BOM Cost (%s)" % cur in ci:
                    op_vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(
                        op.get("bom_cost")
                    )
                    if has_usd and "BOM Cost (%s)" % usd in ci and rate:
                        oc = op.get("bom_cost") or 0
                        op_vals[ci["BOM Cost (%s)" % usd] - 1] = (
                            _flt(oc * rate) if oc else ""
                        )
                _tw_row(ws, row, op_vals, _C["tree_op"],
                        outline=child_outline + 1)
                row += 1

    # ── Subcontracting row for this BOM level ────────────────────────────────
    sc = node.get("subcontracting")
    if sc:
        sc_vals = [""] * col_count
        sc_vals[ci["Name"] - 1] = (
            "    " * child_outline
            + "\u25b6 Subcontracting: " + _str(sc.get("name", ""))
        )
        sc_vals[ci["Qty"] - 1] = _flt(sc.get("quantity"))
        sc_vals[ci["UoM"] - 1] = _str(sc.get("uom", ""))
        if show_costs and "BOM Cost (%s)" % cur in ci:
            sc_vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(sc.get("bom_cost"))
            if has_usd and "BOM Cost (%s)" % usd in ci and rate:
                sc_cost = sc.get("bom_cost") or 0
                sc_vals[ci["BOM Cost (%s)" % usd] - 1] = (
                    _flt(sc_cost * rate) if sc_cost else ""
                )
            if "Product Cost (%s)" % cur in ci:
                sc_vals[ci["Product Cost (%s)" % cur] - 1] = _flt(sc.get("prod_cost"))
        _tw_row(ws, row, sc_vals, _C["tree_sc"], outline=child_outline, bold=True)
        row += 1

    # ── Byproducts group for this BOM level ──────────────────────────────────
    bps = node.get("byproducts", [])
    if bps:
        bp_cost = node.get("byproducts_cost",
                           sum(b.get("bom_cost", 0) for b in bps))
        bp_vals = [""] * col_count
        bp_vals[ci["Name"] - 1] = (
            "    " * child_outline + "\u25b6 Byproducts (%d)" % len(bps)
        )
        if show_costs and "BOM Cost (%s)" % cur in ci:
            bp_vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(bp_cost)
            if has_usd and "BOM Cost (%s)" % usd in ci and rate:
                bp_vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(bp_cost * rate)
        _tw_row(ws, row, bp_vals, _C["tree_bp"],
                outline=child_outline, bold=True)
        row += 1

        for bp in bps:
            bp_item_vals = [""] * col_count
            bp_item_vals[ci["Name"] - 1] = (
                "    " * (child_outline + 1) + _str(bp.get("name", ""))
            )
            bp_item_vals[ci["Qty"] - 1] = _flt(bp.get("quantity"))
            bp_item_vals[ci["UoM"] - 1] = _str(bp.get("uom_name", ""))
            if show_costs:
                _tree_costs(
                    bp_item_vals, ci, cur, usd, has_usd, rate,
                    bp.get("bom_cost"), bp.get("prod_cost"),
                )
            _tw_row(ws, row, bp_item_vals, _C["tree_bp_item"],
                    outline=child_outline + 1)
            row += 1

    return row


def _build_tree_sheet(ws, bom_lines, cur, usd, rate,
                      show_costs, show_operations, show_lead_times,
                      bom_name, quantity):
    """
    Write Sheet 1: BOM hierarchical tree mirroring the Odoo BOM Overview.

    Columns mirror the native UI: Product, Qty, UoM, Lead Time, Route,
    BOM Cost, Product Cost (+ optional USD conversion), Free-to-Use, On Hand.

    Operations and byproducts at each BOM level are written as collapsible
    row groups (Excel outline) immediately below their parent BOM row.
    """
    from openpyxl.styles import Alignment  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    has_usd = bool(usd and rate)

    # ── Column definitions ────────────────────────────────────────────────────
    cols = ["Name", "Qty", "UoM"]
    if show_lead_times:
        cols += ["Lead Time (days)", "Route"]
    if show_costs:
        cols.append("BOM Cost (%s)" % cur)
        if has_usd:
            cols.append("BOM Cost (%s)" % usd)
        cols.append("Product Cost (%s)" % cur)
        if has_usd:
            cols.append("Product Cost (%s)" % usd)
    cols += ["Free to Use", "On Hand", "Availability"]

    col_count = len(cols)
    ci = {name: idx + 1 for idx, name in enumerate(cols)}

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 50   # Name (indented tree)
    ws.column_dimensions["B"].width = 10   # Qty
    ws.column_dimensions["C"].width = 8    # UoM
    col_d_letter = get_column_letter(4)
    ws.column_dimensions[col_d_letter].width = 15  # Lead Time / BOM Cost
    for letter_i in range(5, col_count + 1):
        ws.column_dimensions[get_column_letter(letter_i)].width = 16

    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 1

    # ── BOM info banner ───────────────────────────────────────────────────────
    _write_info(ws, row, "BOM", bom_name, col_count)
    row += 1
    _write_info(ws, row, "Quantity", str(quantity), col_count)
    row += 1
    row += 1  # blank separator

    # ── Column-header row ─────────────────────────────────────────────────────
    header_row = row
    for col_idx, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = _font(bold=True, color=_C["header_fg"])
        c.fill = _fill(_C["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    row += 1

    # ── Tree ─────────────────────────────────────────────────────────────────
    _write_tree_node(
        ws, row, bom_lines, ci, cur, usd, rate,
        has_usd, show_costs, show_operations, show_lead_times,
        outline_level=0,
        col_count=col_count,
    )


# ════════════════════════════ Sheet 2: Cost Summary ══════════════════════════

def _build_summary_sheet(ws, cs, cur, usd,
                         show_costs, show_operations, show_lead_times,
                         bom_name, quantity):
    """
    Write Sheet 1: hierarchical Cost Summary with Excel row groups.

    Row group convention (summaryBelow=False means summary row is ABOVE
    its children, so collapse buttons appear above the grouped block):

      outline_level 0 → always visible (category depth 0, workcenter)
      outline_level 1 → child of level-0 (category depth 1, product of depth-0)
      outline_level 2 → child of level-1 (category depth 2, product of depth-1,
                                          usage under depth-0 product)
      …and so on.
    """
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    has_usd = bool(cs and usd and any(
        w.get("total_usd") for w in cs.get("workcenters", [])
    ) or any(
        n.get("total_usd") for n in cs.get("categories", [])
    ))

    # ── Build ordered column list ─────────────────────────────────────────────
    cols = ["Level", "Type", "Name", "Qty / Duration", "UoM", "%"]
    if show_operations:
        cols += ["Qty (ud)", "min/ud", "ud/hr"]
    if show_costs:
        cols.append("BOM Cost (%s)" % cur)
        if has_usd:
            cols.append("BOM Cost (%s)" % usd)
        cols.append("Product Cost (%s)" % cur)
        if has_usd:
            cols.append("Product Cost (%s)" % usd)
    if show_lead_times:
        cols += ["Lead Time (days)", "Route"]
    cols += ["Free to Use", "On Hand", "Availability"]

    col_count = len(cols)
    # ci: column name → 1-based column index
    ci = {name: idx + 1 for idx, name in enumerate(cols)}

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7    # Level
    ws.column_dimensions["B"].width = 13   # Type
    ws.column_dimensions["C"].width = 46   # Name
    ws.column_dimensions["D"].width = 15   # Qty/Duration
    ws.column_dimensions["E"].width = 8    # UoM
    ws.column_dimensions["F"].width = 8    # %
    for i in range(7, col_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Collapse buttons appear ABOVE their group (not below).
    ws.sheet_properties.outlinePr.summaryBelow = False

    row = 1

    # ── BOM info banner ───────────────────────────────────────────────────────
    _write_info(ws, row, "BOM", bom_name, col_count)
    row += 1
    _write_info(ws, row, "Quantity", str(quantity), col_count)
    row += 1
    row += 1  # blank separator

    # ── Column-header row ─────────────────────────────────────────────────────
    header_row = row
    from openpyxl.styles import Alignment  # noqa: PLC0415
    for col_idx, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = _font(bold=True, color=_C["header_fg"])
        c.fill = _fill(_C["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26

    # ── Column header comments (hover tooltips explaining each formula) ──────
    _hdr_comments = {
        "Level": (
            "Hierarchy depth level.\n"
            "  0 = top-level category or work center\n"
            "  1 = sub-category or product\n"
            "  2+ = deeper nesting (usage / operation)\n"
            "Used only to support grouping/indentation logic."
        ),
        "Type": (
            "Row type identifier:\n"
            "  Category   — product category grouping\n"
            "  Product    — component, aggregated across all its usages\n"
            "  Usage      — single occurrence inside one specific parent product\n"
            "  Workcenter — work center header (Operations section)\n"
            "  Operation  — individual manufacturing operation\n"
            "  Subtotal   — section total\n"
            "  TOTAL      — grand total (Components + Operations)"
        ),
        "Name": (
            "Name of the category, component, work center, or operation.\n"
            "Indented with 4 spaces per hierarchy level."
        ),
        "Qty / Duration": (
            "Components section:\n"
            "  Usage row   — quantity consumed per unit of finished product.\n"
            "  Product row — sum of quantities across all usages\n"
            "               (blank /— when UoMs differ).\n\n"
            "Operations section:\n"
            "  Operation row  — duration in minutes.\n"
            "  Workcenter row — sum of durations of all its operations."
        ),
        "UoM": (
            "Unit of measure for the quantity.\n"
            "'—' = mixed UoMs across usages; aggregation not possible."
        ),
        "%": (
            "Percentage share of this row's BOM Cost within its section:\n\n"
            "Components (Category / Product / Usage):\n"
            "  % = row BOM Cost ÷ Subtotal Components × 100\n\n"
            "Operations (Workcenter / Operation):\n"
            "  % = row BOM Cost ÷ Subtotal Operations × 100\n\n"
            "Subtotal rows:\n"
            "  % = Subtotal ÷ Grand Total × 100"
        ),
        "BOM Cost (%s)" % cur: (
            "Total contribution of this row to the BOM cost.\n\n"
            "  Usage:       qty × product unit cost × BOM scale factor\n"
            "  Product:     Σ BOM Costs of all usages of this component\n"
            "  Category:    Σ BOM Costs of products + child categories\n"
            "  Operation:   (duration ÷ 60) × work center cost/hour\n"
            "  Workcenter:  Σ BOM Costs of all its operations\n"
            "  Subtotal:    Σ BOM Costs of direct rows in this section\n"
            "  Grand Total: Subtotal Components + Subtotal Operations"
        ),
        "Product Cost (%s)" % cur: (
            "Catalogue unit cost × quantity. Does NOT include operations\n"
            "or manufacturing overhead.\n\n"
            "  Usage:    qty × product.standard_price\n"
            "  Product:  Σ Product Costs of all usages\n"
            "  Category: Σ Product Costs of products + child categories\n"
            "  Subtotal: Σ Product Costs of the Components section"
        ),
        "Lead Time (days)": (
            "Supplier or manufacturing lead time in calendar days.\n"
            "Derived from the replenishment route assigned to this\n"
            "component (purchase order lead time, manufacture lead\n"
            "time, resupply lead time, etc.)."
        ),
        "Route": (
            "Replenishment route for this component\n"
            "(e.g. Buy, Manufacture, Resupply from warehouse).\n"
            "May include sub-route detail such as vendor name."
        ),
        "Free to Use": (
            "Available-to-promise quantity.\n"
            "Formula: On Hand − Reserved (outgoing)\n"
            "Source: stock.quant → virtual_available field."
        ),
        "On Hand": (
            "Total physical quantity in stock across all locations.\n"
            "Source: stock.quant → quantity field."
        ),
        "Availability": (
            "Stock availability status vs. required quantity:\n"
            "  Available     — Free to Use ≥ required quantity\n"
            "  Partial       — some stock, but insufficient\n"
            "  Not Available — no usable stock"
        ),
    }
    if has_usd:
        _hdr_comments["BOM Cost (%s)" % usd] = (
            "BOM Cost converted to %s using the company's\n"
            "exchange rate at the time of export.\n"
            "Formula: BOM Cost (%s) × rate(%s→%s)" % (usd, cur, cur, usd)
        )
        _hdr_comments["Product Cost (%s)" % usd] = (
            "Product Cost converted to %s.\n"
            "Formula: Product Cost (%s) × rate(%s→%s)" % (usd, cur, cur, usd)
        )
    for _ci, _col_name in enumerate(cols, start=1):
        if _col_name in _hdr_comments:
            _cell_comment(ws, header_row, _ci, _hdr_comments[_col_name])

    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    row += 1

    # ── Components section ────────────────────────────────────────────────────
    if cs["categories"]:
        _write_section_header(ws, row, "▶  Components by Product Category", col_count)
        row += 1

        for node in cs["categories"]:
            row = _write_category_rows(
                ws, row, node, ci, cur, usd,
                has_usd, show_costs, show_lead_times,
                outline_base=0,
            )

        _sub_comp_row = row
        row = _write_subtotal(
            ws, row, "Subtotal Components", col_count, ci,
            bom_cost=cs["totals"]["components"],
            bom_cost_usd=cs["totals"].get("components_usd"),
            prod_cost=cs["totals"].get("prod_cost"),
            prod_cost_usd=cs["totals"].get("prod_cost_usd"),
            show_costs=show_costs, has_usd=has_usd, cur=cur, usd=usd,
        )
        if show_costs and "BOM Cost (%s)" % cur in ci:
            _cell_comment(
                ws, _sub_comp_row, ci["BOM Cost (%s)" % cur],
                "TOTAL BOM COST — Components\n\n"
                "= Σ (qty_i × product.standard_price × BOM_scale_factor)\n"
                "  for all components at the requested production quantity.\n\n"
                "Includes nested sub-assembly costs recursively.",
            )
        if show_costs and "Product Cost (%s)" % cur in ci:
            _cell_comment(
                ws, _sub_comp_row, ci["Product Cost (%s)" % cur],
                "TOTAL PRODUCT COST — Components\n\n"
                "= Σ (qty_i × product.standard_price)\n"
                "  for all components at the requested production quantity.\n\n"
                "Uses the catalogue standard_price; does NOT include\n"
                "manufacturing operations or overhead.",
            )
        row += 1  # blank separator

    # ── Operations section ────────────────────────────────────────────────────
    if show_operations and cs["workcenters"]:
        _write_section_header(ws, row, "▶  Operations by Work Center", col_count)
        row += 1

        for wc in cs["workcenters"]:
            # Workcenter header row (always visible → outline_level 0)
            vals = [""] * col_count
            vals[ci["Level"] - 1] = 0
            vals[ci["Type"] - 1] = "Workcenter"
            vals[ci["Name"] - 1] = wc["name"]
            vals[ci["Qty / Duration"] - 1] = _flt(wc.get("total_duration"))
            vals[ci["%"] - 1] = _pct(wc.get("percentage"))
            if show_costs:
                vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(wc.get("total"))
                if has_usd and "BOM Cost (%s)" % usd in ci:
                    vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(wc.get("total_usd"))
            _write_row(ws, row, vals, _C["wc"], bold=True, outline_level=0)
            row += 1

            for op in wc.get("items", []):
                op_name = _str(op.get("name"))
                if op.get("parent_name"):
                    op_name += "  ← " + _str(op["parent_name"])
                vals = [""] * col_count
                vals[ci["Level"] - 1] = 1
                vals[ci["Type"] - 1] = "Operation"
                vals[ci["Name"] - 1] = "        " + op_name
                vals[ci["Qty / Duration"] - 1] = _flt(op.get("duration"))
                vals[ci["%"] - 1] = _pct(op.get("percentage"))
                if show_operations:
                    _parent_qty = op.get("parent_qty") or 1
                    _duration = op.get("duration") or 0
                    vals[ci["Qty (ud)"] - 1] = _flt(_parent_qty)
                    vals[ci["min/ud"] - 1] = (
                        round(_duration / _parent_qty, 2)
                        if _duration and _parent_qty
                        else ""
                    )
                    vals[ci["ud/hr"] - 1] = (
                        round((_parent_qty * 60) / _duration, 2)
                        if _duration and _parent_qty
                        else ""
                    )
                if show_costs:
                    vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(op.get("total"))
                    if has_usd and "BOM Cost (%s)" % usd in ci:
                        vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(op.get("total_usd"))
                if show_lead_times:
                    lt = op.get("lead_time")
                    vals[ci["Lead Time (days)"] - 1] = lt if lt is not False and lt is not None else ""
                    vals[ci["Route"] - 1] = _str(op.get("route_name"))
                _write_row(ws, row, vals, _C["op"], outline_level=1)
                row += 1

        _sub_ops_row = row
        row = _write_subtotal(
            ws, row, "Subtotal Operations", col_count, ci,
            bom_cost=cs["totals"]["operations"],
            bom_cost_usd=cs["totals"].get("operations_usd"),
            prod_cost=None,
            prod_cost_usd=None,
            show_costs=show_costs, has_usd=has_usd, cur=cur, usd=usd,
        )
        if show_costs and "BOM Cost (%s)" % cur in ci:
            _cell_comment(
                ws, _sub_ops_row, ci["BOM Cost (%s)" % cur],
                "TOTAL BOM COST — Operations\n\n"
                "= Σ ((duration_j ÷ 60) × work_center_cost_per_hour)\n"
                "  for all manufacturing operations in this BOM.\n\n"
                "Duration is in minutes; cost rate is the work center's\n"
                "time efficiency + capacity × cost/hour setting.",
            )
        row += 1  # blank separator

    # ── Subcontracting by Vendor section ────────────────────────────────────
    sc_vendors = cs.get("subcontracting", [])
    if sc_vendors:
        _write_section_header(ws, row, "▶  Subcontracting by Vendor", col_count)
        row += 1

        for vendor in sc_vendors:
            # Vendor group row (outline_level 0)
            vals = [""] * col_count
            vals[ci["Type"] - 1] = "Vendor"
            vals[ci["Name"] - 1] = vendor["name"]
            vals[ci["%"] - 1] = _pct(vendor.get("percentage"))
            if show_costs:
                vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(vendor.get("total"))
                if has_usd and "BOM Cost (%s)" % usd in ci:
                    vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(vendor.get("total_usd"))
                if "Product Cost (%s)" % cur in ci:
                    vals[ci["Product Cost (%s)" % cur] - 1] = _flt(vendor.get("prod_cost_total"))
                if has_usd and "Product Cost (%s)" % usd in ci:
                    vals[ci["Product Cost (%s)" % usd] - 1] = _flt(vendor.get("prod_cost_total_usd"))
            _write_row(ws, row, vals, _C["sc_vendor"], bold=True, outline_level=0)
            row += 1

            for sc_item in vendor.get("items", []):
                vals = [""] * col_count
                vals[ci["Level"] - 1] = 1
                vals[ci["Type"] - 1] = "Subcontracting"
                vals[ci["Name"] - 1] = "        " + _str(sc_item.get("product_name"))
                vals[ci["Qty / Duration"] - 1] = _flt(sc_item.get("quantity"))
                vals[ci["%"] - 1] = _pct(sc_item.get("percentage"))
                if show_costs:
                    vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(sc_item.get("total"))
                    if has_usd and "BOM Cost (%s)" % usd in ci:
                        vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(sc_item.get("total_usd"))
                    if "Product Cost (%s)" % cur in ci:
                        vals[ci["Product Cost (%s)" % cur] - 1] = _flt(sc_item.get("prod_cost"))
                    if has_usd and "Product Cost (%s)" % usd in ci:
                        vals[ci["Product Cost (%s)" % usd] - 1] = _flt(sc_item.get("prod_cost_usd"))
                _write_row(ws, row, vals, _C["sc_item"], outline_level=1)
                row += 1

        # Subtotal Subcontracting
        _sub_sc_row = row
        row = _write_subtotal(
            ws, row, "Subtotal Subcontracting", col_count, ci,
            bom_cost=cs["totals"]["subcontracting"],
            bom_cost_usd=cs["totals"].get("subcontracting_usd"),
            prod_cost=cs["totals"].get("subcontracting_prod_cost"),
            prod_cost_usd=cs["totals"].get("subcontracting_prod_cost_usd"),
            show_costs=show_costs, has_usd=has_usd, cur=cur, usd=usd,
            color=_C["sc_subtotal"],
        )
        row += 1  # blank separator

    # ── Byproducts by Category section ──────────────────────────────────────
    bp_categories = cs.get("byproductCategories", [])
    if bp_categories:
        _write_section_header(ws, row, "▶  Byproducts by Category", col_count)
        row += 1

        for bp_node in bp_categories:
            row = _write_byproduct_category_rows(
                ws, row, bp_node, ci, cur, usd,
                has_usd, show_costs,
                outline_base=0,
            )

        # Subtotal Byproducts — written directly to use the green colour.
        _sub_bp_row = row
        sub_vals = [""] * col_count
        sub_vals[ci["Type"] - 1] = "Subtotal"
        sub_vals[ci["Name"] - 1] = "Subtotal Byproducts"
        if show_costs:
            sub_vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(
                cs["totals"]["byproducts"]
            )
            if has_usd and "BOM Cost (%s)" % usd in ci:
                sub_vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(
                    cs["totals"].get("byproducts_usd")
                )
            if "Product Cost (%s)" % cur in ci:
                sub_vals[ci["Product Cost (%s)" % cur] - 1] = _flt(
                    cs["totals"].get("byproducts_prod_cost")
                )
            if has_usd and "Product Cost (%s)" % usd in ci:
                sub_vals[ci["Product Cost (%s)" % usd] - 1] = _flt(
                    cs["totals"].get("byproducts_prod_cost_usd")
                )
        _write_row(ws, _sub_bp_row, sub_vals, _C["bp_subtotal"], bold=True)
        if show_costs and "BOM Cost (%s)" % cur in ci:
            _cell_comment(
                ws, _sub_bp_row, ci["BOM Cost (%s)" % cur],
                "SUBTOTAL BYPRODUCTS — BOM Cost (Allocated Away)\n\n"
                "= Σ (qty_k × std_cost_k × cost_share_factor)\n"
                "  for all byproducts.\n\n"
                "Represents the portion of BOM cost allocated to byproducts\n"
                "via cost_share. May be 0 when cost_share = 0%.",
            )
        if show_costs and "Product Cost (%s)" % cur in ci:
            _cell_comment(
                ws, _sub_bp_row, ci["Product Cost (%s)" % cur],
                "SUBTOTAL BYPRODUCTS — Recoverable Value\n\n"
                "= Σ (qty_k × product.standard_price)  for all byproducts.\n\n"
                "Catalogue value of co-products / recovered materials.\n"
                "Always meaningful, independent of cost_share.",
            )
        row += 1  # blank separator

    # ── Grand Total ───────────────────────────────────────────────────────────
    # Row 1: TOTAL (Gross — Components + Operations) — BoM Cost and Prod Cost together.
    vals = [""] * col_count
    vals[ci["Type"] - 1] = "TOTAL"
    vals[ci["Name"] - 1] = "Total  (Components + Operations + Subcontracting)"
    if show_costs:
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(cs["totals"]["total"])
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(cs["totals"].get("total_usd"))
        if "Product Cost (%s)" % cur in ci:
            vals[ci["Product Cost (%s)" % cur] - 1] = _flt(cs["totals"].get("total_prod"))
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(cs["totals"].get("total_prod_usd"))
    _write_row(ws, row, vals, _C["total"], bold=True)
    ws.row_dimensions[row].height = 15
    if show_costs and "BOM Cost (%s)" % cur in ci:
        _cell_comment(
            ws, row, ci["BOM Cost (%s)" % cur],
            "TOTAL BOM COST (Gross)\n\n"
            "= Subtotal Components + Subtotal Operations\n\n"
            "Components: \u03a3 (qty_i \u00d7 std_cost_i \u00d7 production_qty_factor)\n"
            "Operations: \u03a3 ((duration_j \u00f7 60) \u00d7 wc_cost_per_hour)\n\n"
            "Gross value before deducting byproduct recoverable value.\n"
            "Scaled to the production quantity shown in the BOM header.",
        )
    total_row = row
    row += 1

    # Rows 2 & 3: only when byproducts exist
    if bp_categories and show_costs and "Product Cost (%s)" % cur in ci:
        # Row 2: (−) RECOVERABLE BYPRODUCT VALUE — both BoM and Prod columns
        vals = [""] * col_count
        vals[ci["Type"] - 1] = "TOTAL"
        vals[ci["Name"] - 1] = "(\u2212) Recoverable byproduct value"
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(cs["totals"].get("byproducts"))
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(cs["totals"].get("byproducts_usd"))
        vals[ci["Product Cost (%s)" % cur] - 1] = _flt(
            cs["totals"].get("byproducts_prod_cost")
        )
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(
                cs["totals"].get("byproducts_prod_cost_usd")
            )
        _write_row(ws, row, vals, _C["bp_subtotal"], bold=True)
        ws.row_dimensions[row].height = 15
        _cell_comment(
            ws, row, ci["Product Cost (%s)" % cur],
            "BYPRODUCTS RECOVERABLE VALUE\n\n"
            "BOM Cost: \u03a3 (qty_k \u00d7 std_cost_k \u00d7 cost_share_k)  for all byproducts.\n"
            "Prod Cost: \u03a3 (qty_k \u00d7 product.standard_price)  for all byproducts.\n\n"
            "Catalogue value of recovered co-products, subtracted to obtain the NET COST.",
        )
        row += 1

        # Row 3: (=) NET COST — both BoM and Prod columns
        vals = [""] * col_count
        vals[ci["Type"] - 1] = "TOTAL"
        vals[ci["Name"] - 1] = "(=) NET COST"
        vals[ci["BOM Cost (%s)" % cur] - 1] = _flt(cs["totals"].get("net_bom"))
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = _flt(cs["totals"].get("net_bom_usd"))
        vals[ci["Product Cost (%s)" % cur] - 1] = _flt(
            cs["totals"].get("net_prod")
        )
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = _flt(
                cs["totals"].get("net_prod_usd")
            )
        _write_row(ws, row, vals, _C["total"], bold=True)
        ws.row_dimensions[row].height = 15
        _cell_comment(
            ws, row, ci["Product Cost (%s)" % cur],
            "NET COST\n\n"
            "= Total \u2212 Byproducts Recoverable Value\n\n"
            "Effective cost of the finished product after accounting for\n"
            "the value recovered from byproducts / co-products.\n"
            "Can be negative when co-product value exceeds input cost.",
        )
        row += 1

    # ── Auto-filter on header row ─────────────────────────────────────────────
    # Use row - 1 so the filter covers all written rows including any
    # additional grand-total rows added after the TOTAL BOM COST row.
    ws.auto_filter.ref = (
        "A%d:%s%d" % (header_row, get_column_letter(col_count), row - 1)
    )


# ════════════════════════════ Sheet 2: Components Detail ═════════════════════

def _flatten_usages(cs):
    """
    Walk the category tree and yield one dict per component usage.

    :returns generator of dicts with keys:
        category_path, product_name, parent_name,
        quantity, uom_name, percentage,
        bom_cost, bom_cost_usd, prod_cost, prod_cost_usd,
        lead_time, route_name, route_detail, route_type,
        quantity_available, quantity_on_hand, availability_state, availability_display
    """
    def _walk(node, path_parts):
        path_parts = path_parts + [node["name"]]
        cat_path = " > ".join(path_parts)

        # Recurse into children first so rows are ordered depth-first
        for child in node.get("children", []):
            yield from _walk(child, path_parts)

        for prod in node.get("products", []):
            for usage in prod.get("usages", []):
                yield {
                    "category_path":       cat_path,
                    "product_name":        prod.get("name", ""),
                    "parent_name":         _str(usage.get("parent_name")),
                    "quantity":            _flt(usage.get("quantity")),
                    "uom_name":            _str(usage.get("uom_name")),
                    "percentage":          _pct(usage.get("percentage")),
                    "bom_cost":            _flt(usage.get("total")),
                    "bom_cost_usd":        _flt(usage.get("total_usd")),
                    "prod_cost":           _flt(usage.get("prod_cost")),
                    "prod_cost_usd":       _flt(usage.get("prod_cost_usd")),
                    "lead_time":           usage.get("lead_time"),
                    "route_name":          _str(usage.get("route_name")),
                    "route_detail":        _str(usage.get("route_detail")),
                    "route_type":          _str(usage.get("route_type")),
                    "quantity_available":  prod.get("quantity_available"),
                    "quantity_on_hand":    prod.get("quantity_on_hand"),
                    "availability_display": _str(prod.get("availability_display")),
                }

    for root_node in cs.get("categories", []):
        yield from _walk(root_node, [])


def _build_detail_sheet(ws, cs, cur, usd, show_lead_times):
    """
    Write Sheet 2: flat, pivot-ready component usage list.
    One data row per usage (component × parent-product pair).
    """
    from openpyxl.styles import Alignment  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    has_usd = bool(usd and any(
        n.get("total_usd") for n in cs.get("categories", [])
    ))

    # ── Column definitions ────────────────────────────────────────────────────
    cols = [
        "Category",
        "Product",
        "Used In (Parent)",
        "Quantity",
        "UoM",
        "% of Components",
        "BOM Cost (%s)" % cur,
    ]
    if has_usd:
        cols.append("BOM Cost (%s)" % usd)
    cols.append("Product Cost (%s)" % cur)
    if has_usd:
        cols.append("Product Cost (%s)" % usd)
    if show_lead_times:
        cols += ["Lead Time (days)", "Route", "Route Detail", "Route Type"]
    cols += ["Free to Use", "On Hand", "Availability"]

    col_count = len(cols)
    ci = {name: idx + 1 for idx, name in enumerate(cols)}

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    for i in range(7, col_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = 1
    for col_idx, h in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = _font(bold=True, color=_C["header_fg"])
        c.fill = _fill(_C["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    # ── Column header comments (hover tooltips) ───────────────────────────────
    _det_comments = {
        "Category": (
            "Full category path for this component, from root to leaf.\n"
            "Format: Root > Sub-category > ... > Leaf category"
        ),
        "Product": "Component product name.",
        "Used In (Parent)": (
            "Parent product or sub-assembly that directly uses\n"
            "this component in its bill of materials."
        ),
        "Quantity": (
            "Quantity of this component consumed per unit of\n"
            "the finished product (or parent sub-assembly).\n"
            "Scaled to the production quantity of the BOM export."
        ),
        "UoM": "Unit of measure for the quantity.",
        "% of Components": (
            "Percentage share of this usage's BOM Cost\n"
            "relative to the total Components cost.\n\n"
            "Formula: usage BOM Cost ÷ Subtotal Components × 100"
        ),
        "BOM Cost (%s)" % cur: (
            "Cost contribution of this usage to the BOM.\n\n"
            "Formula: qty × product unit cost × BOM scale factor"
        ),
        "Product Cost (%s)" % cur: (
            "Catalogue unit cost × quantity.\n\n"
            "Formula: qty × product.standard_price\n"
            "Does NOT include operations or overhead."
        ),
        "Lead Time (days)": (
            "Supplier or manufacturing lead time in calendar days,\n"
            "from the replenishment route of this component."
        ),
        "Route": (
            "Replenishment route (e.g. Buy, Manufacture, MTO).\n"
            "May include sub-route detail such as vendor name."
        ),
        "Route Detail": "Additional route detail (e.g. vendor name or sub-route).",
        "Route Type": "Route type code (buy / manufacture / resupply / push / pull).",
        "Free to Use": (
            "Available-to-promise quantity.\n"
            "Formula: On Hand − Reserved\n"
            "Source: stock.quant → virtual_available."
        ),
        "On Hand": (
            "Total physical quantity in stock.\n"
            "Source: stock.quant → quantity."
        ),
        "Availability": (
            "Stock availability vs. required quantity:\n"
            "  Available     — Free to Use ≥ required quantity\n"
            "  Partial       — some stock, but insufficient\n"
            "  Not Available — no usable stock"
        ),
    }
    if has_usd:
        _det_comments["BOM Cost (%s)" % usd] = (
            "BOM Cost converted to %s.\n"
            "Formula: BOM Cost (%s) × rate(%s→%s)" % (usd, cur, cur, usd)
        )
        _det_comments["Product Cost (%s)" % usd] = (
            "Product Cost converted to %s.\n"
            "Formula: Product Cost (%s) × rate(%s→%s)" % (usd, cur, cur, usd)
        )
    for _ci2, _col_name2 in enumerate(cols, start=1):
        if _col_name2 in _det_comments:
            _cell_comment(ws, header_row, _ci2, _det_comments[_col_name2])

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = header_row + 1
    for item in _flatten_usages(cs):
        vals = [""] * col_count
        vals[ci["Category"] - 1] = item["category_path"]
        vals[ci["Product"] - 1] = item["product_name"]
        vals[ci["Used In (Parent)"] - 1] = item["parent_name"]
        vals[ci["Quantity"] - 1] = item["quantity"]
        vals[ci["UoM"] - 1] = item["uom_name"]
        vals[ci["% of Components"] - 1] = item["percentage"]
        vals[ci["BOM Cost (%s)" % cur] - 1] = item["bom_cost"]
        if has_usd and "BOM Cost (%s)" % usd in ci:
            vals[ci["BOM Cost (%s)" % usd] - 1] = item["bom_cost_usd"]
        vals[ci["Product Cost (%s)" % cur] - 1] = item["prod_cost"]
        if has_usd and "Product Cost (%s)" % usd in ci:
            vals[ci["Product Cost (%s)" % usd] - 1] = item["prod_cost_usd"]
        if show_lead_times:
            lt = item.get("lead_time")
            vals[ci["Lead Time (days)"] - 1] = lt if lt is not False and lt is not None else ""
            vals[ci["Route"] - 1] = item["route_name"]
            vals[ci["Route Detail"] - 1] = item["route_detail"]
            vals[ci["Route Type"] - 1] = item["route_type"]
        qa = item.get("quantity_available")
        if qa is not False and qa is not None and "Free to Use" in ci:
            vals[ci["Free to Use"] - 1] = _flt(qa)
            vals[ci["On Hand"] - 1] = _flt(item.get("quantity_on_hand"))
        vals[ci["Availability"] - 1] = item["availability_display"]

        _write_row(ws, row, vals, _C["usage"])
        row += 1

    # ── Auto-filter ───────────────────────────────────────────────────────────
    if row > header_row + 1:
        ws.auto_filter.ref = (
            "A%d:%s%d" % (header_row, get_column_letter(col_count), row - 1)
        )


# ════════════════════════════ Controller ═════════════════════════════════════

class BomCostSummaryXlsxController(http.Controller):

    @http.route(
        "/econovo/bom_cost_summary/export_xlsx",
        type="http",
        auth="user",
    )
    def export_xlsx(
        self,
        bom_id,
        quantity=None,
        variant=None,
        warehouse_id=None,
        costs="true",
        operations="true",
        lead_times="true",
        **_kw,
    ):
        """Generate and return a .xlsx for the BOM Cost Summary."""
        try:
            from openpyxl import Workbook  # noqa: PLC0415
        except ImportError:
            _logger.error(
                "econovo_mrp_bom_cost_summary: openpyxl is not installed; "
                "cannot generate XLSX export."
            )
            return request.make_response(
                "openpyxl is required for Excel export.", status=500
            )

        # ── Validate bom_id ───────────────────────────────────────────────────
        try:
            bom_id = int(bom_id)
        except (TypeError, ValueError):
            return request.make_response("Invalid bom_id.", status=400)

        bom = request.env["mrp.bom"].browse(bom_id)
        if not bom.exists():
            return request.make_response("BOM not found.", status=404)

        # ── Parse parameters ──────────────────────────────────────────────────
        qty = (
            float(quantity)
            if quantity and quantity not in ("false", "null")
            else (bom.product_qty or 1.0)
        )
        show_costs = str(costs).lower() not in ("false", "0")
        show_operations = str(operations).lower() not in ("false", "0")
        show_lead_times = str(lead_times).lower() not in ("false", "0")

        # ── Fetch BOM data ────────────────────────────────────────────────────
        bom_report = request.env["report.mrp.report_bom_structure"]
        if warehouse_id and warehouse_id not in ("false", "null"):
            bom_report = bom_report.with_context(warehouse=int(warehouse_id))

        variant_id = (
            int(variant)
            if variant and variant not in ("false", "null", "0")
            else False
        )
        raw = bom_report.get_html(
            bom_id=bom_id,
            searchQty=qty,
            searchVariant=variant_id,
        )
        bom_lines = raw.get("lines", {})
        secondary = raw.get("secondary_currency", False)
        rate = secondary.get("rate", 0) if secondary else 0

        report_model = request.env[
            "report.econovo_mrp_bom_cost_summary.report_cost_summary"
        ]
        # Reuse the summary computed server-side and attached by
        # ReportBomStructure._get_report_data (single source of truth shared
        # with the interactive UI and the PDF). Fall back only if missing.
        cost_summary = bom_lines.get("cost_summary")
        if cost_summary is None:
            cost_summary = report_model._compute_cost_summary(
                bom_lines, secondary
            )
        if not cost_summary:
            return request.make_response(
                "No cost data available for this BOM.", status=204
            )

        currency_name = bom.company_id.currency_id.name
        usd_name = secondary.get("name", "USD") if secondary else ""

        # ── Build workbook ────────────────────────────────────────────────────
        wb = Workbook()

        # Sheet 1: BOM Tree (first sheet — active when opening)
        ws0 = wb.active
        ws0.title = "BOM Tree"
        _build_tree_sheet(
            ws0, bom_lines, currency_name, usd_name, rate,
            show_costs, show_operations, show_lead_times,
            bom.display_name, qty,
        )

        # Sheet 2: Cost Summary (by category / work center)
        ws1 = wb.create_sheet("Cost Summary")
        _build_summary_sheet(
            ws1, cost_summary, currency_name, usd_name,
            show_costs, show_operations, show_lead_times,
            bom.display_name, qty,
        )

        # Sheet 3: Components Detail (flat pivot-ready list)
        ws2 = wb.create_sheet("Components Detail")
        _build_detail_sheet(
            ws2, cost_summary, currency_name, usd_name, show_lead_times,
        )

        # ── Serialize ─────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        xlsx_data = output.getvalue()

        safe_name = (
            bom.display_name
            .replace("/", "-")
            .replace("\\", "-")
            .replace(" ", "_")[:60]
        )
        filename = "BOM_Cost_Summary_%s.xlsx" % safe_name

        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet",
            ),
            ("Content-Disposition", content_disposition(filename)),
        ]
        return request.make_response(xlsx_data, headers=headers)
