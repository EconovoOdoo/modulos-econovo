# -*- coding: utf-8 -*-
"""HTTP controller: export MO Cost Summary to Excel (.xlsx).

Endpoint: GET /econovo/mo_cost_summary/export_xlsx
Params:   production_id (required)

Returns a .xlsx workbook with up to four sheets:
  Sheet 1 "MO Tree"                 — hierarchical MO breakdown mirroring the
                                       native Odoo MO Overview tree view, with
                                       expandable operations per level and
                                       sub-MOs shown as child nodes.
  Sheet 2 "Components by Category"  — MO cost / Real cost grouped by product category.
  Sheet 3 "Operations by Work Center" — MO cost / Real cost grouped by work center.
  Sheet 4 "Subcontracting by Vendor" — (only when subcontracting sub-MOs are present)
                                        Subcontracting cost grouped by vendor.
"""

import io
import logging

from odoo import http
from odoo.http import content_disposition, request

_logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
_C = {
    # General
    "header_bg":        "1F4E79",
    "header_fg":        "FFFFFF",
    # Components by Category sheet
    "cat_0":            "BDD7EE",
    "cat_1":            "DDEBF7",
    "cat_deep":         "EEF4FB",
    "product":          "F2F2F2",
    "usage":            "FFFFFF",
    # Operations sheet
    "wc":               "E2EFDA",
    "op":               "F4F9F5",
    # Totals
    "subtotal":         "FFF2CC",
    "total":            "FFE082",
    # MO Tree sheet
    "tree_root":        "1F4E79",   # root MO row (dark blue)
    "tree_root_fg":     "FFFFFF",   # root MO foreground (white)
    "tree_mo":          "BDD7EE",   # sub-MO rows (medium blue)
    "tree_comp_mo":     "DDEBF7",   # component row that has sub-MO replenishments
    "tree_leaf":        "F2F2F2",   # leaf component rows (light gray)
    "tree_ops_grp":     "E2EFDA",   # operations group header (green)
    "tree_op":          "F4F9F5",   # individual operation rows (light green)
    "tree_sc":          "E8D5F5",   # subcontracting sub-MO rows (purple)
    # Subcontracting sheet
    "sc_vendor":        "E8D5F5",   # vendor group rows (purple)
    "sc_item":          "F5EEF8",   # subcontracting item rows (lighter purple)
    "sc_total":         "D7BDE2",   # subcontracting grand total row
    # Byproducts sheet (green tones — value recovered from the process)
    "bp_cat_0":         "C6E0B4",   # byproduct category depth 0
    "bp_cat_1":         "D9EAD3",   # byproduct category depth 1
    "bp_cat_deep":      "EAF4E7",   # byproduct category depth 2+
    "bp_product":       "F3FAF1",   # byproduct product rows
    "bp_total":         "A9D18E",   # byproducts grand total row
}


def _fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=9):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size)


def _align_right():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="right", vertical="center")


def _align_left():
    from openpyxl.styles import Alignment
    return Alignment(horizontal="left", vertical="center")


def _flt(value):
    if value is None or value is False:
        return ""
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return ""


def _str(value):
    """Return value as string, empty string for None/False."""
    if value is None or value is False:
        return ""
    return str(value)


def _write_row(ws, row_idx, values, bg_color, bold=False, outline_level=0):
    fill = _fill(bg_color)
    row_font = _font(bold=bold)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = row_font
        cell.fill = fill
        if isinstance(val, (int, float)) and col_idx > 3:
            cell.alignment = _align_right()
        else:
            cell.alignment = _align_left()
    if outline_level:
        ws.row_dimensions[row_idx].outline_level = min(outline_level, 7)


def _tw_row(ws, row_idx, vals, bg, outline=0, bold=False, fg="000000"):
    """Write a tree row with custom foreground colour support."""
    fnt = _font(bold=bold, color=fg)
    fill = _fill(bg)
    for col_idx, val in enumerate(vals, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = fnt
        cell.fill = fill
        if isinstance(val, (int, float)) and col_idx > 2:
            cell.alignment = _align_right()
        else:
            cell.alignment = _align_left()
    if outline:
        ws.row_dimensions[row_idx].outline_level = min(outline, 7)


def _write_header(ws, row_idx, values, bg_color):
    fill = _fill(bg_color)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = _font(bold=True, color=_C["header_fg"], size=9)
        cell.fill = fill
        cell.alignment = _align_left()


# ══════════════════════════ Sheet 1: MO Tree ═════════════════════════════════

def _collect_sub_mo_ids(comp_wrappers, out_ids):
    """Recursively collect all sub-MO production IDs from replenishments."""
    for wrapper in comp_wrappers:
        for rep in (wrapper.get("replenishments") or []):
            rep_sum = rep.get("summary") or {}
            if rep_sum.get("model") == "mrp.production":
                pid = rep_sum.get("id")
                if pid:
                    out_ids.add(pid)
                _collect_sub_mo_ids(rep.get("components") or [], out_ids)


def _write_mo_tree_level(ws, row, name, qty, uom_name, mo_cost, real_cost,
                          comp_wrappers, operations, wc_resolver,
                          subcontractor_map, ci, cur, col_count, outline_level,
                          is_root=False, is_subcontract=False):
    """Recursively write one level of the MO component tree.

    Parameters
    ----------
    ws                : openpyxl worksheet
    row               : current 1-based row index
    name              : product / sub-MO display name
    qty / uom_name    : quantity and unit of measure
    mo_cost / real_cost : costs for this level's product row
    comp_wrappers     : list of component wrapper dicts (each has summary +
                        replenishments); children of this level
    operations        : the operations dict with 'details' key for this level
    wc_resolver       : callable(index, op_id) → {workcenter_id, workcenter_name, …}
    subcontractor_map : {production_id: vendor_name} for subcontracting sub-MOs
    ci                : column-index dict, e.g. {"Name": 1, "Qty": 2, …}
    cur               : currency name string (e.g. "ARS")
    col_count         : total number of columns
    outline_level     : current Excel outline depth (0 = root)
    is_root           : True only for the top-level MO product row
    is_subcontract    : True when this sub-MO is a subcontracting order
    """
    # Choose row styling
    if is_root:
        bg, fg, bold = _C["tree_root"], _C["tree_root_fg"], True
    elif is_subcontract:
        bg, fg, bold = _C["tree_sc"], "000000", True
    else:
        bg, fg, bold = _C["tree_mo"], "000000", True

    indent = "    " * outline_level
    vals = [""] * col_count
    vals[ci["Name"] - 1] = indent + name
    vals[ci["Qty"] - 1] = _flt(qty)
    vals[ci["UoM"] - 1] = _str(uom_name)
    vals[ci["MO Cost (%s)" % cur] - 1] = _flt(mo_cost)
    vals[ci["Real Cost (%s)" % cur] - 1] = _flt(real_cost)
    if mo_cost:
        vals[ci["Deviation"] - 1] = _flt(real_cost - mo_cost)
    _tw_row(ws, row, vals, bg, outline=outline_level, bold=bold, fg=fg)
    row += 1

    # Operations group for this level
    op_details = (operations or {}).get("details") or []
    if op_details:
        total_op_mo = sum(float(op.get("mo_cost") or 0) for op in op_details)
        total_op_real = sum(float(op.get("real_cost") or 0) for op in op_details)
        grp_indent = "    " * (outline_level + 1)
        grp_vals = [""] * col_count
        grp_vals[ci["Name"] - 1] = grp_indent + "▶ Operations (%d)" % len(op_details)
        grp_vals[ci["MO Cost (%s)" % cur] - 1] = _flt(total_op_mo)
        grp_vals[ci["Real Cost (%s)" % cur] - 1] = _flt(total_op_real)
        if total_op_mo:
            grp_vals[ci["Deviation"] - 1] = _flt(total_op_real - total_op_mo)
        _write_row(ws, row, grp_vals, _C["tree_ops_grp"], bold=True,
                   outline_level=outline_level + 1)
        row += 1
        for i, op in enumerate(op_details):
            extra = wc_resolver(i, op.get("id") or 0)
            wc_name = extra.get("workcenter_name") or ""
            op_name = op.get("name") or ""
            if wc_name:
                op_name = wc_name + ": " + op_name
            op_indent = "    " * (outline_level + 2)
            op_vals = [""] * col_count
            op_vals[ci["Name"] - 1] = op_indent + op_name
            op_vals[ci["Qty"] - 1] = _flt(op.get("quantity"))
            op_vals[ci["UoM"] - 1] = "min"
            op_vals[ci["MO Cost (%s)" % cur] - 1] = _flt(op.get("mo_cost"))
            op_vals[ci["Real Cost (%s)" % cur] - 1] = _flt(op.get("real_cost"))
            _write_row(ws, row, op_vals, _C["tree_op"],
                       outline_level=outline_level + 2)
            row += 1

    # Children: component wrappers
    for wrapper in (comp_wrappers or []):
        comp_sum = wrapper.get("summary") or {}
        comp_name = comp_sum.get("name") or ""
        comp_qty = comp_sum.get("quantity")
        comp_uom = comp_sum.get("uom_name") or ""
        comp_mo = float(comp_sum.get("mo_cost") or 0)
        comp_real = float(comp_sum.get("real_cost") or 0)

        sub_mo_reps = [
            rep for rep in (wrapper.get("replenishments") or [])
            if (rep.get("summary") or {}).get("model") == "mrp.production"
        ]

        c_indent = "    " * (outline_level + 1)

        if sub_mo_reps:
            # Component is manufactured → write as an intermediate node,
            # then write each sub-MO as its child.
            c_vals = [""] * col_count
            c_vals[ci["Name"] - 1] = c_indent + comp_name
            c_vals[ci["Qty"] - 1] = _flt(comp_qty)
            c_vals[ci["UoM"] - 1] = _str(comp_uom)
            c_vals[ci["MO Cost (%s)" % cur] - 1] = _flt(comp_mo)
            c_vals[ci["Real Cost (%s)" % cur] - 1] = _flt(comp_real)
            if comp_mo:
                c_vals[ci["Deviation"] - 1] = _flt(comp_real - comp_mo)
            _write_row(ws, row, c_vals, _C["tree_comp_mo"], bold=True,
                       outline_level=outline_level + 1)
            row += 1

            for rep in sub_mo_reps:
                rep_sum = rep.get("summary") or {}
                rep_id = rep_sum.get("id") or 0
                is_sc = rep_id in subcontractor_map
                sc_vendor = subcontractor_map.get(rep_id, "")
                sub_name = rep_sum.get("name") or comp_name
                if is_sc and sc_vendor:
                    display_name = "Subcontract [%s]: %s" % (sc_vendor, sub_name)
                elif is_sc:
                    display_name = "Subcontract: %s" % sub_name
                else:
                    display_name = "Sub-MO: %s" % sub_name
                rep_wc_map = rep.get("operations_workcenter_map") or {}
                sub_resolver = (
                    lambda i, op_id, _m=rep_wc_map: _m.get(op_id) or {}
                )
                row = _write_mo_tree_level(
                    ws, row,
                    name=display_name,
                    qty=rep_sum.get("quantity"),
                    uom_name=rep_sum.get("uom_name") or comp_uom,
                    mo_cost=float(rep_sum.get("mo_cost") or 0),
                    real_cost=float(rep_sum.get("real_cost") or 0),
                    comp_wrappers=rep.get("components") or [],
                    operations=rep.get("operations"),
                    wc_resolver=sub_resolver,
                    subcontractor_map=subcontractor_map,
                    ci=ci, cur=cur, col_count=col_count,
                    outline_level=outline_level + 2,
                    is_root=False,
                    is_subcontract=is_sc,
                )
        else:
            # Leaf component → single row
            c_vals = [""] * col_count
            c_vals[ci["Name"] - 1] = c_indent + comp_name
            c_vals[ci["Qty"] - 1] = _flt(comp_qty)
            c_vals[ci["UoM"] - 1] = _str(comp_uom)
            c_vals[ci["MO Cost (%s)" % cur] - 1] = _flt(comp_mo)
            c_vals[ci["Real Cost (%s)" % cur] - 1] = _flt(comp_real)
            if comp_mo:
                c_vals[ci["Deviation"] - 1] = _flt(comp_real - comp_mo)
            _write_row(ws, row, c_vals, _C["tree_leaf"],
                       outline_level=outline_level + 1)
            row += 1

    return row


def _write_mo_tree_sheet(ws, data, currency_name, subcontractor_map):
    """Write Sheet 1: MO Tree — hierarchical component breakdown."""
    cur = currency_name
    headers = [
        "Name", "Qty", "UoM",
        "MO Cost (%s)" % cur,
        "Real Cost (%s)" % cur,
        "Deviation",
    ]
    ci = {h: i + 1 for i, h in enumerate(headers)}
    col_count = len(headers)

    _write_header(ws, 1, headers, _C["header_bg"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.sheet_properties.outlinePr.summaryBelow = False

    summary = data.get("summary") or {}
    prod_name = summary.get("name") or "MO"
    prod_qty = summary.get("quantity")
    prod_uom = summary.get("uom_name") or ""
    prod_mo_cost = float(summary.get("mo_cost") or 0)
    prod_real_cost = float(summary.get("real_cost") or 0)

    top_wc_info = data.get("operations_workcenter_info") or []
    root_resolver = (
        lambda i, op_id, _info=top_wc_info: _info[i] if i < len(_info) else {}
    )

    _write_mo_tree_level(
        ws, 2,
        name=prod_name,
        qty=prod_qty,
        uom_name=prod_uom,
        mo_cost=prod_mo_cost,
        real_cost=prod_real_cost,
        comp_wrappers=data.get("components") or [],
        operations=data.get("operations"),
        wc_resolver=root_resolver,
        subcontractor_map=subcontractor_map,
        ci=ci, cur=cur, col_count=col_count,
        outline_level=0,
        is_root=True,
    )


# ══════════════════════ Sheet 4: Subcontracting by Vendor ════════════════════

def _collect_subcontracting_entries(comp_wrappers, subcontractor_map, out):
    """Recursively collect subcontracting replenishment entries."""
    for wrapper in comp_wrappers:
        comp_sum = wrapper.get("summary") or {}
        comp_name = comp_sum.get("name") or ""
        for rep in (wrapper.get("replenishments") or []):
            rep_sum = rep.get("summary") or {}
            if rep_sum.get("model") == "mrp.production":
                rep_id = rep_sum.get("id") or 0
                if rep_id in subcontractor_map:
                    out.append({
                        "vendor_name":     subcontractor_map[rep_id],
                        "product_name":    rep_sum.get("name") or comp_name,
                        "parent_name":     comp_name,
                        "quantity":        rep_sum.get("quantity"),
                        "uom_name":        rep_sum.get("uom_name") or "",
                        "mo_cost":         float(rep_sum.get("mo_cost") or 0),
                        "real_cost":       float(rep_sum.get("real_cost") or 0),
                    })
                # Recurse into this sub-MO's components regardless
                _collect_subcontracting_entries(
                    rep.get("components") or [], subcontractor_map, out
                )


def _write_subcontracting_sheet(ws, data, currency_name, subcontractor_map):
    """Write Sheet 4: Subcontracting by Vendor (only when subcontracting exists)."""
    sc_entries = []
    _collect_subcontracting_entries(
        data.get("components") or [], subcontractor_map, sc_entries
    )
    if not sc_entries:
        return False

    cur = currency_name
    headers = [
        "Type", "Name", "Parent Component", "Qty", "UoM",
        "MO Cost (%s)" % cur,
        "Real Cost (%s)" % cur,
        "Deviation",
    ]
    _write_header(ws, 1, headers, _C["header_bg"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 14

    # Group by vendor
    vendor_map = {}
    for entry in sc_entries:
        vn = entry["vendor_name"]
        if vn not in vendor_map:
            vendor_map[vn] = {"mo_cost": 0.0, "real_cost": 0.0, "items": []}
        vendor_map[vn]["mo_cost"] += entry["mo_cost"]
        vendor_map[vn]["real_cost"] += entry["real_cost"]
        vendor_map[vn]["items"].append(entry)

    row = 2
    total_mo = 0.0
    total_real = 0.0
    for vendor_name, vdata in sorted(vendor_map.items()):
        v_dev = _flt(vdata["real_cost"] - vdata["mo_cost"]) if vdata["mo_cost"] else ""
        _write_row(ws, row, [
            "Vendor", vendor_name, "", "", "",
            _flt(vdata["mo_cost"]), _flt(vdata["real_cost"]), v_dev,
        ], _C["sc_vendor"], bold=True, outline_level=0)
        row += 1
        total_mo += vdata["mo_cost"]
        total_real += vdata["real_cost"]

        for item in vdata["items"]:
            i_dev = _flt(item["real_cost"] - item["mo_cost"]) if item["mo_cost"] else ""
            _write_row(ws, row, [
                "Subcontracting",
                "  " + item["product_name"],
                item["parent_name"],
                _flt(item["quantity"]),
                item["uom_name"],
                _flt(item["mo_cost"]),
                _flt(item["real_cost"]),
                i_dev,
            ], _C["sc_item"], outline_level=1)
            row += 1

    # Grand total row
    if vendor_map:
        t_dev = _flt(total_real - total_mo) if total_mo else ""
        _write_row(ws, row, [
            "TOTAL", "", "", "", "",
            _flt(total_mo), _flt(total_real), t_dev,
        ], _C["sc_total"], bold=True)

    return True


# ── Sheet 2: Components by Category ──────────────────────────────────────────

def _collect_component_entries(comp_wrappers, parent_name, out):
    """Recursively collect flat component entries from the MO report tree.

    Each wrapper produced by _get_components_data has:
      - wrapper["summary"]: the native component dict (product_id, name,
        quantity, uom_name, mo_cost, real_cost, …)
      - wrapper["categ_id"], wrapper["categ_name"], wrapper["categ_ancestors"]:
        injected by our Python override to avoid polluting MoOverviewLine's
        strict OWL prop-shape validation on "summary".
      - wrapper["replenishments"]: list of replenishment dicts, where those
        with summary.model == "mrp.production" are sub-MOs carrying their own
        "components" list.

    The resulting flat entries are keyed with "_parent_name" so the sheet
    can show which MO consumed each component.
    """
    for wrapper in comp_wrappers:
        summary = wrapper.get("summary") or {}
        entry = dict(summary)
        # categ fields live on the wrapper, not on the summary
        entry["categ_id"] = wrapper["categ_id"] if "categ_id" in wrapper else entry.get("categ_id", 0)
        entry["categ_name"] = wrapper.get("categ_name") or entry.get("categ_name", "Uncategorized")
        entry["categ_ancestors"] = wrapper.get("categ_ancestors") or entry.get("categ_ancestors") or []
        entry["_parent_name"] = parent_name
        out.append(entry)

        # Recurse into sub-MO replenishments (each carries its own components list).
        for rep in (wrapper.get("replenishments") or []):
            rep_sum = rep.get("summary") or {}
            if rep_sum.get("model") == "mrp.production":
                sub_comps = rep.get("components") or []
                sub_name = rep_sum.get("name") or parent_name
                _collect_component_entries(sub_comps, sub_name, out)


def _build_category_map(flat_entries):
    """Aggregate a flat component entry list into {categ_id: {...}}."""
    category_map = {}
    for entry in flat_entries:
        cat_id = entry.get("categ_id") or 0
        cat_name = entry.get("categ_name") or "Uncategorized"
        ancestors = entry.get("categ_ancestors") or [{"id": cat_id, "name": cat_name}]

        if cat_id not in category_map:
            category_map[cat_id] = {
                "id": cat_id,
                "name": cat_name,
                "ancestors": ancestors,
                "mo_cost": 0.0,
                "real_cost": 0.0,
                "products": {},
            }
        cat = category_map[cat_id]
        mo_cost = float(entry.get("mo_cost") or 0)
        real_cost = float(entry.get("real_cost") or 0)
        cat["mo_cost"] += mo_cost
        cat["real_cost"] += real_cost

        prod_id = entry.get("product_id") or entry.get("id") or 0
        prod_name = entry.get("name") or ""
        if prod_id not in cat["products"]:
            cat["products"][prod_id] = {
                "name": prod_name,
                "mo_cost": 0.0,
                "real_cost": 0.0,
                "usages": [],
            }
        prod = cat["products"][prod_id]
        prod["mo_cost"] += mo_cost
        prod["real_cost"] += real_cost
        prod["usages"].append({
            "parent_name": entry.get("_parent_name") or "",
            "quantity": float(entry.get("quantity") or 0),
            "uom_name": entry.get("uom_name") or "",
            "mo_cost": mo_cost,
            "real_cost": real_cost,
        })
    return category_map


def _write_components_sheet(ws, data, currency_name):
    """Write Sheet 1: Components by Category."""
    # data["summary"] holds the MO-level summary; fall back to data["name"].
    parent_name = (data.get("summary") or {}).get("name") or data.get("name", "")
    flat_entries = []
    _collect_component_entries(data.get("components") or [], parent_name, flat_entries)
    category_map = _build_category_map(flat_entries)

    headers = ["Type", "Name", "Qty", "UoM",
               "MO Cost (%s)" % currency_name,
               "Real Cost (%s)" % currency_name,
               "Deviation"]
    _write_header(ws, 1, headers, _C["header_bg"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    row = 2
    total_mo = 0.0
    total_real = 0.0

    for cat in sorted(category_map.values(), key=lambda x: -x["mo_cost"]):
        dev = _flt(cat["real_cost"] - cat["mo_cost"]) if cat["mo_cost"] else ""
        _write_row(ws, row, [
            "Category", cat["name"], "", "",
            _flt(cat["mo_cost"]), _flt(cat["real_cost"]), dev,
        ], _C["cat_0"], bold=True, outline_level=0)
        row += 1
        total_mo += cat["mo_cost"]
        total_real += cat["real_cost"]

        for prod in sorted(cat["products"].values(), key=lambda p: -p["mo_cost"]):
            dev_p = _flt(prod["real_cost"] - prod["mo_cost"]) if prod["mo_cost"] else ""
            _write_row(ws, row, [
                "Product", "  " + prod["name"], "", "",
                _flt(prod["mo_cost"]), _flt(prod["real_cost"]), dev_p,
            ], _C["product"], bold=False, outline_level=1)
            row += 1

            for usage in prod["usages"]:
                _write_row(ws, row, [
                    "Usage",
                    "    " + usage["parent_name"],
                    _flt(usage["quantity"]),
                    usage["uom_name"],
                    _flt(usage["mo_cost"]),
                    _flt(usage["real_cost"]),
                    _flt(usage["real_cost"] - usage["mo_cost"]) if usage["mo_cost"] else "",
                ], _C["usage"], bold=False, outline_level=2)
                row += 1

    _write_row(ws, row, [
        "TOTAL", "", "", "",
        _flt(total_mo), _flt(total_real),
        _flt(total_real - total_mo) if total_mo else "",
    ], _C["total"], bold=True)


# ── Sheet 2: Operations by Work Center ───────────────────────────────────────

def _collect_operation_entries(comp_wrappers, out):
    """Recursively collect operation entries from sub-MO replenishments.

    For each component wrapper, inspect its replenishments for sub-MOs and
    collect their workorders using the operations_workcenter_map stored at
    the replenishment level (our Python override moves it there from inside
    operations to avoid OWL prop-shape violations in MoOverviewComponentsBlock).

    Recurses into each sub-MO's own components to collect deeper levels.
    """
    for wrapper in comp_wrappers:
        summary = wrapper.get("summary") or {}
        for rep in (wrapper.get("replenishments") or []):
            rep_sum = rep.get("summary") or {}
            if rep_sum.get("model") != "mrp.production":
                continue
            ops_details = (rep.get("operations") or {}).get("details") or []
            # workcenter_map is at the replenishment level, not inside operations
            wc_map_rep = rep.get("operations_workcenter_map") or {}
            for op in ops_details:
                wc_info = wc_map_rep.get(op.get("id")) or {}
                out.append({
                    "name": op.get("name") or "",
                    "quantity": float(op.get("quantity") or 0),
                    "mo_cost": float(op.get("mo_cost") or 0),
                    "real_cost": float(op.get("real_cost") or 0),
                    "workcenter_id": wc_info.get("workcenter_id") or 0,
                    "workcenter_name": wc_info.get("workcenter_name") or "Unknown",
                })
            # Recurse into this sub-MO's own component wrappers for deeper levels.
            _collect_operation_entries(rep.get("components") or [], out)


def _write_operations_sheet(ws, data, currency_name):
    """Write Sheet 2: Operations by Work Center."""
    headers = ["Type", "Name", "Duration (min)",
               "MO Cost (%s)" % currency_name,
               "Real Cost (%s)" % currency_name,
               "Deviation"]
    _write_header(ws, 1, headers, _C["header_bg"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    # ── Top-level operations ──────────────────────────────────────────────────
    # operations_workcenter_info is indexed in the same order as operations.details
    # (our _get_report_data override moves workcenter_map to this sibling key).
    top_details = (data.get("operations") or {}).get("details") or []
    top_wc_info = data.get("operations_workcenter_info") or []
    top_ops = []
    for i, op in enumerate(top_details):
        extra = top_wc_info[i] if i < len(top_wc_info) else {}
        top_ops.append({
            "name": op.get("name") or "",
            "quantity": float(op.get("quantity") or 0),
            "mo_cost": float(op.get("mo_cost") or 0),
            "real_cost": float(op.get("real_cost") or 0),
            "workcenter_id": extra.get("workcenter_id") or 0,
            "workcenter_name": extra.get("workcenter_name") or "Unknown",
        })

    # ── Sub-MO operations ─────────────────────────────────────────────────────
    sub_ops = []
    _collect_operation_entries(data.get("components") or [], sub_ops)

    # ── Group all operations by workcenter ────────────────────────────────────
    wc_map = {}
    for op in top_ops + sub_ops:
        wc_id = op["workcenter_id"]
        if wc_id not in wc_map:
            wc_map[wc_id] = {
                "name": op["workcenter_name"],
                "mo_cost": 0.0,
                "real_cost": 0.0,
                "duration": 0.0,
                "items": [],
            }
        wc = wc_map[wc_id]
        wc["mo_cost"] += op["mo_cost"]
        wc["real_cost"] += op["real_cost"]
        wc["duration"] += op["quantity"]
        wc["items"].append(op)

    row = 2
    total_mo = 0.0
    total_real = 0.0

    for wc in sorted(wc_map.values(), key=lambda x: -x["mo_cost"]):
        _write_row(ws, row, [
            "Work Center", wc["name"], _flt(wc["duration"]),
            _flt(wc["mo_cost"]), _flt(wc["real_cost"]),
            _flt(wc["real_cost"] - wc["mo_cost"]) if wc["mo_cost"] else "",
        ], _C["wc"], bold=True, outline_level=0)
        row += 1
        total_mo += wc["mo_cost"]
        total_real += wc["real_cost"]

        for op in wc["items"]:
            _write_row(ws, row, [
                "Operation", "  " + op["name"], _flt(op["quantity"]),
                _flt(op["mo_cost"]), _flt(op["real_cost"]),
                _flt(op["real_cost"] - op["mo_cost"]) if op["mo_cost"] else "",
            ], _C["op"], bold=False, outline_level=1)
            row += 1

    if wc_map:
        _write_row(ws, row, [
            "TOTAL", "", "",
            _flt(total_mo), _flt(total_real),
            _flt(total_real - total_mo) if total_mo else "",
        ], _C["total"], bold=True)


# ── Sheet 4: Byproducts by Category ──────────────────────────────────────────

def _collect_byproduct_entries(data, parent_name, out):
    """Collect byproduct entries from the MO report data tree.

    Reads ``data["byproducts"]["details"]`` for the current level and then
    recurses into sub-MO replenishments so that byproducts at every
    manufacturing level are included.

    Category info is stored in ``data["byproducts"]["categ_map"]`` keyed by
    product_id (not injected directly into the detail dicts, which are passed
    to MoOverviewLine and have a strict OWL prop shape).

    Each entry dict contains the byproduct fields (name, quantity, uom_name,
    mo_cost, real_cost, categ_id, categ_name, categ_ancestors) plus a
    ``_parent_name`` key indicating which MO produced it.
    """
    byproducts_data = data.get("byproducts") or {}
    categ_map = byproducts_data.get("categ_map") or {}
    for bp in byproducts_data.get("details") or []:
        entry = dict(bp)
        entry["_parent_name"] = parent_name
        categ_info = categ_map.get(bp.get("id")) or {}
        entry.update(categ_info)
        out.append(entry)

    # Recurse into sub-MO replenishments embedded in component wrappers.
    for wrapper in (data.get("components") or []):
        for rep in (wrapper.get("replenishments") or []):
            rep_sum = rep.get("summary") or {}
            if rep_sum.get("model") == "mrp.production":
                sub_name = rep_sum.get("name") or parent_name
                _collect_byproduct_entries(rep, sub_name, out)


def _build_byproduct_category_map(flat_entries):
    """Aggregate flat byproduct entries into {categ_id: {...}}."""
    category_map = {}
    for entry in flat_entries:
        cat_id = entry.get("categ_id") or 0
        cat_name = entry.get("categ_name") or "Uncategorized"
        ancestors = entry.get("categ_ancestors") or [{"id": cat_id, "name": cat_name}]

        if cat_id not in category_map:
            category_map[cat_id] = {
                "id": cat_id,
                "name": cat_name,
                "ancestors": ancestors,
                "mo_cost": 0.0,
                "real_cost": 0.0,
                "products": {},
            }
        cat = category_map[cat_id]
        mo_cost = float(entry.get("mo_cost") or 0)
        real_cost = float(entry.get("real_cost") or 0)
        cat["mo_cost"] += mo_cost
        cat["real_cost"] += real_cost

        prod_id = entry.get("id") or 0
        prod_name = entry.get("name") or ""
        if prod_id not in cat["products"]:
            cat["products"][prod_id] = {
                "name": prod_name,
                "mo_cost": 0.0,
                "real_cost": 0.0,
                "usages": [],
            }
        prod = cat["products"][prod_id]
        prod["mo_cost"] += mo_cost
        prod["real_cost"] += real_cost
        prod["usages"].append({
            "parent_name": entry.get("_parent_name") or "",
            "quantity": float(entry.get("quantity") or 0),
            "uom_name": entry.get("uom_name") or "",
            "mo_cost": mo_cost,
            "real_cost": real_cost,
        })
    return category_map


def _write_byproducts_sheet(ws, data, currency_name):
    """Write the Byproducts by Category sheet.

    Returns True when at least one byproduct was written, False otherwise.
    """
    parent_name = (data.get("summary") or {}).get("name") or data.get("name", "")
    flat_entries = []
    _collect_byproduct_entries(data, parent_name, flat_entries)
    if not flat_entries:
        return False

    category_map = _build_byproduct_category_map(flat_entries)

    cur = currency_name
    headers = ["Type", "Name", "Qty", "UoM",
               "MO Cost (%s)" % cur,
               "Real Cost (%s)" % cur,
               "Deviation"]
    _write_header(ws, 1, headers, _C["header_bg"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    cat_colors = [_C["bp_cat_0"], _C["bp_cat_1"], _C["bp_cat_deep"]]

    row = 2
    total_mo = 0.0
    total_real = 0.0

    for cat in sorted(category_map.values(), key=lambda x: -x["mo_cost"]):
        cat_color = cat_colors[0]
        dev = _flt(cat["real_cost"] - cat["mo_cost"]) if cat["mo_cost"] else ""
        _write_row(ws, row, [
            "BP Category", cat["name"], "", "",
            _flt(cat["mo_cost"]), _flt(cat["real_cost"]), dev,
        ], cat_color, bold=True, outline_level=0)
        row += 1
        total_mo += cat["mo_cost"]
        total_real += cat["real_cost"]

        for prod in sorted(cat["products"].values(), key=lambda p: -p["mo_cost"]):
            dev_p = _flt(prod["real_cost"] - prod["mo_cost"]) if prod["mo_cost"] else ""
            _write_row(ws, row, [
                "BP Product", "  " + prod["name"], "", "",
                _flt(prod["mo_cost"]), _flt(prod["real_cost"]), dev_p,
            ], _C["bp_product"], bold=False, outline_level=1)
            row += 1

            for usage in prod["usages"]:
                dev_u = _flt(usage["real_cost"] - usage["mo_cost"]) if usage["mo_cost"] else ""
                _write_row(ws, row, [
                    "BP Usage",
                    "    " + usage["parent_name"],
                    _flt(usage["quantity"]),
                    usage["uom_name"],
                    _flt(usage["mo_cost"]),
                    _flt(usage["real_cost"]),
                    dev_u,
                ], _C["usage"], bold=False, outline_level=2)
                row += 1

    _write_row(ws, row, [
        "TOTAL", "", "", "",
        _flt(total_mo), _flt(total_real),
        _flt(total_real - total_mo) if total_mo else "",
    ], _C["bp_total"], bold=True)

    return True


# ── Controller ────────────────────────────────────────────────────────────────

class MoCostSummaryXlsx(http.Controller):

    @http.route(
        "/econovo/mo_cost_summary/export_xlsx",
        type="http",
        auth="user",
        methods=["GET"],
    )
    def export_xlsx(self, production_id=None, **kwargs):
        """Export the MO Cost Summary as an .xlsx workbook."""
        try:
            from openpyxl import Workbook
        except ImportError:
            return request.make_response(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl",
                headers=[("Content-Type", "text/plain")],
            )

        if not production_id:
            return request.make_response(
                "Missing production_id parameter",
                headers=[("Content-Type", "text/plain")],
            )

        try:
            prod_id = int(production_id)
        except (TypeError, ValueError):
            return request.make_response(
                "Invalid production_id",
                headers=[("Content-Type", "text/plain")],
            )

        env = request.env
        report = env["report.mrp.report_mo_overview"]
        result = report.get_report_values(prod_id)
        data = result["data"]

        # Determine currency name (mrp.production has no direct currency_id;
        # the currency comes from the production order's company).
        production = env["mrp.production"].browse(prod_id)
        currency_name = production.company_id.currency_id.name or "ARS"

        # Build subcontractor map: {production_id: vendor_display_name}
        # by querying all sub-MO production records for subcontractor_id.
        sub_mo_ids = set()
        _collect_sub_mo_ids(data.get("components") or [], sub_mo_ids)
        subcontractor_map = {}
        if sub_mo_ids:
            sub_productions = env["mrp.production"].browse(list(sub_mo_ids))
            for sub_prod in sub_productions:
                if sub_prod.subcontractor_id:
                    subcontractor_map[sub_prod.id] = (
                        sub_prod.subcontractor_id.display_name
                    )

        # Build workbook
        wb = Workbook()

        # Sheet 1: MO Tree
        ws1 = wb.active
        ws1.title = "MO Tree"
        _write_mo_tree_sheet(ws1, data, currency_name, subcontractor_map)

        # Sheet 2: Components by Category
        ws2 = wb.create_sheet("Components by Category")
        _write_components_sheet(ws2, data, currency_name)

        # Sheet 3: Operations by Work Center
        ws3 = wb.create_sheet("Operations by Work Center")
        _write_operations_sheet(ws3, data, currency_name)

        # Sheet 4: Subcontracting by Vendor (only when subcontracting exists)
        if subcontractor_map:
            ws4 = wb.create_sheet("Subcontracting by Vendor")
            _write_subcontracting_sheet(ws4, data, currency_name, subcontractor_map)

        # Sheet 5: Byproducts by Category (only when byproducts exist)
        ws_bp = wb.create_sheet("Byproducts by Category")
        if not _write_byproducts_sheet(ws_bp, data, currency_name):
            wb.remove(ws_bp)

        # Stream as .xlsx download
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = "mo_cost_summary_%d.xlsx" % prod_id
        headers = [
            ("Content-Type",
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ("Content-Disposition", content_disposition(filename)),
        ]
        return request.make_response(buf.getvalue(), headers=headers)
